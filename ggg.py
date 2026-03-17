"""
================================================================================
  ENTERPRISE-GRADE DATA QUALITY (DQ) FRAMEWORK
  Financial Data Validation: Source vs Target
  Author  : Senior Data Engineer
  Version : 2.0.0
  Output  : DQ_Report.xlsx  (multi-sheet, stakeholder-ready)
================================================================================
"""

import warnings
warnings.filterwarnings("ignore")

import io, os, gc, re, hashlib, logging
from datetime import datetime
from typing import Optional

import numpy  as np
import pandas as pd
from scipy  import stats as scipy_stats

from openpyxl                        import Workbook
from openpyxl.styles                 import (Font, PatternFill, Alignment,
                                              Border, Side, GradientFill)
from openpyxl.styles.numbers         import FORMAT_PERCENTAGE_00
from openpyxl.formatting.rule        import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils                  import get_column_letter
from openpyxl.chart                  import BarChart, LineChart, Reference
from openpyxl.chart.series           import DataPoint

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 ─ ROBUST DATA INGESTION
# ═══════════════════════════════════════════════════════════════════════════════

def _clean_bytes(raw: bytes) -> bytes:
    """Strip null bytes and other common binary artefacts."""
    return raw.replace(b"\x00", b"").replace(b"\r\n", b"\n")


def _sniff_delimiter(sample: str) -> str:
    """Auto-detect CSV delimiter from a text sample."""
    for delim in [",", "|", "\t", ";"]:
        if delim in sample:
            return delim
    return ","


def load_csv(
    filepath: str,
    chunksize: int = 500_000,
    key_col: Optional[str] = None,
) -> pd.DataFrame:
    """
    Production-grade CSV loader.
    • Strips null bytes before parsing
    • Auto-detects delimiter
    • Uses chunked reading for memory efficiency
    • Coerces known null-like strings
    """
    log.info(f"Loading: {filepath}")

    # --- read raw bytes, strip null bytes ---
    with open(filepath, "rb") as fh:
        raw = fh.read()
    cleaned = _clean_bytes(raw)
    buf = io.StringIO(cleaned.decode("utf-8", errors="replace"))

    # --- sniff delimiter ---
    sample = buf.read(4096)
    delim  = _sniff_delimiter(sample)
    buf.seek(0)

    NULL_VALS = {"", "NULL", "null", "None", "none", "NA", "N/A", "NaN", "nan", "#N/A"}

    chunks = []
    try:
        reader = pd.read_csv(
            buf,
            sep=delim,
            chunksize=chunksize,
            low_memory=False,
            on_bad_lines="warn",
            na_values=NULL_VALS,
            keep_default_na=True,
            encoding_errors="replace",
            quotechar='"',
            escapechar="\\",
        )
        for chunk in reader:
            chunks.append(chunk)
            gc.collect()
    except Exception as exc:
        log.error(f"Chunk read failed ({exc}), retrying as single read …")
        buf.seek(0)
        return pd.read_csv(buf, sep=delim, low_memory=False, na_values=NULL_VALS)

    df = pd.concat(chunks, ignore_index=True)
    log.info(f"  → {len(df):,} rows × {len(df.columns)} cols loaded")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 ─ COLUMN-TYPE INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════════

_DATE_PATTERNS = [
    r"\d{4}-\d{2}-\d{2}",
    r"\d{2}/\d{2}/\d{4}",
    r"\d{2}-\d{2}-\d{4}",
    r"\d{4}/\d{2}/\d{2}",
]

def detect_column_types(df: pd.DataFrame) -> dict:
    """
    Returns dict mapping col → 'numeric' | 'date' | 'categorical' | 'text'
    Samples up to 500 non-null values per column for speed.
    """
    result = {}
    for col in df.columns:
        s = df[col].dropna()
        sample = s.head(500)

        if pd.api.types.is_numeric_dtype(s):
            result[col] = "numeric"
            continue

        if pd.api.types.is_datetime64_any_dtype(s):
            result[col] = "date"
            continue

        # attempt date parse
        str_sample = sample.astype(str)
        is_date = any(
            str_sample.str.match(pat).mean() > 0.8
            for pat in _DATE_PATTERNS
        )
        if is_date:
            result[col] = "date"
            continue

        # attempt numeric
        try:
            pd.to_numeric(sample, errors="raise")
            result[col] = "numeric"
            continue
        except (ValueError, TypeError):
            pass

        n_unique = s.nunique()
        result[col] = "categorical" if n_unique / max(len(s), 1) < 0.5 else "text"

    return result


def coerce_types(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    """Coerce columns to their detected types in place (copy)."""
    df = df.copy()
    for col, ctype in col_types.items():
        if col not in df.columns:
            continue
        if ctype == "numeric":
            df[col] = pd.to_numeric(df[col], errors="coerce")
        elif ctype == "date":
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 ─ SCHEMA VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def compare_schemas(src: pd.DataFrame, tgt: pd.DataFrame) -> pd.DataFrame:
    """
    Returns a DataFrame with per-column schema comparison:
    presence, data-type match, position shift.
    """
    src_cols = {c: str(src[c].dtype) for c in src.columns}
    tgt_cols = {c: str(tgt[c].dtype) for c in tgt.columns}
    all_cols = sorted(set(src_cols) | set(tgt_cols))

    rows = []
    for col in all_cols:
        in_src   = col in src_cols
        in_tgt   = col in tgt_cols
        src_type = src_cols.get(col, "—")
        tgt_type = tgt_cols.get(col, "—")
        type_match = (src_type == tgt_type) if (in_src and in_tgt) else False

        status = "✅ Both"
        if not in_src:   status = "❌ Missing in Source"
        elif not in_tgt: status = "❌ Missing in Target"
        elif not type_match: status = "⚠️ Type Mismatch"

        src_pos = list(src.columns).index(col) + 1 if in_src else None
        tgt_pos = list(tgt.columns).index(col) + 1 if in_tgt else None

        rows.append({
            "Column"           : col,
            "Status"           : status,
            "Source Dtype"     : src_type,
            "Target Dtype"     : tgt_type,
            "Type Match"       : "Yes" if type_match else "No",
            "Source Position"  : src_pos,
            "Target Position"  : tgt_pos,
            "Position Delta"   : (tgt_pos - src_pos) if (src_pos and tgt_pos) else None,
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 ─ COLUMN-LEVEL PROFILING
# ═══════════════════════════════════════════════════════════════════════════════

def profile_column(series: pd.Series, col_type: str, label: str) -> dict:
    """Compute rich statistics for a single series."""
    n        = len(series)
    null_n   = int(series.isna().sum())
    notnull  = series.dropna()
    nn_n     = len(notnull)
    distinct = int(series.nunique(dropna=True))
    dup_n    = int(nn_n - distinct)

    base = {
        f"{label} Total Rows"    : n,
        f"{label} Null Count"    : null_n,
        f"{label} Null %"        : round(null_n / n * 100, 2) if n else 0,
        f"{label} Non-Null Count": nn_n,
        f"{label} Distinct Count": distinct,
        f"{label} Duplicate Count": max(dup_n, 0),
        f"{label} Fill Rate %"   : round(nn_n / n * 100, 2) if n else 0,
    }

    if col_type == "numeric":
        base.update({
            f"{label} Min"    : float(notnull.min())   if nn_n else None,
            f"{label} Max"    : float(notnull.max())   if nn_n else None,
            f"{label} Mean"   : round(float(notnull.mean()),   4) if nn_n else None,
            f"{label} Median" : round(float(notnull.median()), 4) if nn_n else None,
            f"{label} Std Dev": round(float(notnull.std()),    4) if nn_n else None,
            f"{label} Skew"   : round(float(notnull.skew()),   4) if nn_n else None,
            f"{label} Q1"     : round(float(notnull.quantile(.25)), 4) if nn_n else None,
            f"{label} Q3"     : round(float(notnull.quantile(.75)), 4) if nn_n else None,
        })
    elif col_type == "date":
        base.update({
            f"{label} Min Date" : str(notnull.min()) if nn_n else None,
            f"{label} Max Date" : str(notnull.max()) if nn_n else None,
            f"{label} Date Range (days)": (
                (notnull.max() - notnull.min()).days if nn_n > 1 else 0
            ),
        })
    else:
        top = notnull.value_counts().head(1)
        base.update({
            f"{label} Top Value"       : str(top.index[0])  if nn_n else None,
            f"{label} Top Value Count" : int(top.iloc[0])   if nn_n else None,
            f"{label} Avg String Len"  : round(notnull.astype(str).str.len().mean(), 1) if nn_n else None,
        })
    return base


def profile_all_columns(
    src: pd.DataFrame,
    tgt: pd.DataFrame,
    col_types_src: dict,
    col_types_tgt: dict,
) -> pd.DataFrame:
    """Profile every common column and return a wide comparison DataFrame."""
    common = [c for c in src.columns if c in tgt.columns]
    rows = []

    for col in common:
        ctype = col_types_src.get(col, "text")
        src_stats = profile_column(src[col], ctype, "Source")
        tgt_stats = profile_column(tgt[col], col_types_tgt.get(col, ctype), "Target")

        row = {"Column": col, "Detected Type": ctype}
        row.update(src_stats)
        row.update(tgt_stats)

        # delta metrics for numeric
        if ctype == "numeric":
            for metric in ["Mean", "Median", "Std Dev"]:
                sv = src_stats.get(f"Source {metric}")
                tv = tgt_stats.get(f"Target {metric}")
                if sv is not None and tv is not None:
                    row[f"Δ {metric}"] = round(tv - sv, 4)
                    row[f"Δ {metric} %"] = round((tv - sv) / sv * 100, 2) if sv else None

        # fill-rate delta
        row["Fill Rate Δ %"] = round(
            tgt_stats.get("Target Fill Rate %", 0) -
            src_stats.get("Source Fill Rate %", 0), 2
        )
        rows.append(row)

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 ─ CELL-LEVEL COMPARISON
# ═══════════════════════════════════════════════════════════════════════════════

def _row_hash(df: pd.DataFrame) -> pd.Series:
    """Fast row-level hash using pandas vectorization."""
    return pd.util.hash_pandas_object(df, index=False)


def cell_level_comparison(
    src: pd.DataFrame,
    tgt: pd.DataFrame,
    sample_n: int = 50,
    key_col: Optional[str] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      mismatch_summary  — per-column mismatch counts / %
      mismatch_samples  — up to sample_n example differing rows
    """
    common = [c for c in src.columns if c in tgt.columns]
    min_rows = min(len(src), len(tgt))

    src_cmp = src[common].iloc[:min_rows].reset_index(drop=True)
    tgt_cmp = tgt[common].iloc[:min_rows].reset_index(drop=True)

    summary_rows = []
    sample_rows  = []

    for col in common:
        sv = src_cmp[col].astype(str).str.strip()
        tv = tgt_cmp[col].astype(str).str.strip()
        mismatch_mask = sv != tv
        mm_count = int(mismatch_mask.sum())
        mm_pct   = round(mm_count / min_rows * 100, 4) if min_rows else 0

        summary_rows.append({
            "Column"           : col,
            "Rows Compared"    : min_rows,
            "Mismatch Count"   : mm_count,
            "Mismatch %"       : mm_pct,
            "Match Count"      : min_rows - mm_count,
            "Match %"          : round(100 - mm_pct, 4),
            "DQ Status"        : (
                "✅ Clean" if mm_pct == 0
                else "⚠️ Minor" if mm_pct < 5
                else "🔴 High Mismatch"
            ),
        })

        # collect samples
        if mm_count > 0 and len(sample_rows) < sample_n * 5:
            idx = mismatch_mask[mismatch_mask].index[:sample_n]
            for i in idx:
                sample_rows.append({
                    "Row Index"      : int(i),
                    "Column"         : col,
                    "Source Value"   : src_cmp[col].iloc[i],
                    "Target Value"   : tgt_cmp[col].iloc[i],
                    "Mismatch Type"  : _classify_mismatch(
                        src_cmp[col].iloc[i], tgt_cmp[col].iloc[i]
                    ),
                })

    return pd.DataFrame(summary_rows), pd.DataFrame(sample_rows)


def _classify_mismatch(sv, tv) -> str:
    """Classify the nature of a cell-level mismatch."""
    if pd.isna(sv) and pd.isna(tv): return "Both Null"
    if pd.isna(sv):  return "Null in Source"
    if pd.isna(tv):  return "Null in Target"
    if str(sv).strip().lower() == str(tv).strip().lower():
        return "Case / Whitespace Diff"
    try:
        if abs(float(sv) - float(tv)) < 1e-9: return "Float Precision Diff"
    except (ValueError, TypeError):
        pass
    return "Value Mismatch"


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 ─ ROW-LEVEL CHECKS
# ═══════════════════════════════════════════════════════════════════════════════

def row_level_checks(
    src: pd.DataFrame,
    tgt: pd.DataFrame,
    key_col: Optional[str] = None,
) -> dict:
    """
    Returns a dict with:
      row_counts, duplicate analysis, key-based missing rows.
    """
    result = {}

    # --- counts ---
    result["source_rows"]  = len(src)
    result["target_rows"]  = len(tgt)
    result["row_delta"]    = len(tgt) - len(src)
    result["row_delta_pct"]= round((len(tgt) - len(src)) / max(len(src), 1) * 100, 4)

    # --- duplicates ---
    src_dup = int(src.duplicated().sum())
    tgt_dup = int(tgt.duplicated().sum())
    result["source_duplicates"] = src_dup
    result["target_duplicates"] = tgt_dup
    result["source_dup_pct"]    = round(src_dup / len(src) * 100, 4)
    result["target_dup_pct"]    = round(tgt_dup / len(tgt) * 100, 4)

    # --- key-based missing rows ---
    if key_col and key_col in src.columns and key_col in tgt.columns:
        src_keys = set(src[key_col].dropna().astype(str))
        tgt_keys = set(tgt[key_col].dropna().astype(str))
        result["keys_in_src_not_tgt"]     = len(src_keys - tgt_keys)
        result["keys_in_tgt_not_src"]     = len(tgt_keys - src_keys)
        result["keys_in_both"]            = len(src_keys & tgt_keys)
        result["sample_missing_in_tgt"]   = list(src_keys - tgt_keys)[:20]
        result["sample_missing_in_src"]   = list(tgt_keys - src_keys)[:20]
    else:
        # hash-based row match
        min_r  = min(len(src), len(tgt))
        common = [c for c in src.columns if c in tgt.columns]
        src_h  = _row_hash(src[common].iloc[:min_r])
        tgt_h  = _row_hash(tgt[common].iloc[:min_r])
        matched = int((src_h.values == tgt_h.values).sum())
        result["row_hash_matches"]   = matched
        result["row_hash_mismatches"]= min_r - matched
        result["row_hash_match_pct"] = round(matched / min_r * 100, 4) if min_r else 0

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 ─ TIME-BASED ANALYSIS  (Finance Focus)
# ═══════════════════════════════════════════════════════════════════════════════

def _find_date_col(df: pd.DataFrame, col_types: dict) -> Optional[str]:
    date_cols = [c for c, t in col_types.items() if t == "date" and c in df.columns]
    if not date_cols:
        return None
    # prefer columns with "date", "dt", "mth", "yr" in name
    for hint in ["date", "dt", "mth", "month", "yr", "year", "period"]:
        for dc in date_cols:
            if hint in dc.lower():
                return dc
    return date_cols[0]


def _find_numeric_cols(df: pd.DataFrame, col_types: dict, max_cols: int = 20) -> list:
    return [c for c, t in col_types.items()
            if t == "numeric" and c in df.columns][:max_cols]


def time_based_analysis(
    src: pd.DataFrame,
    tgt: pd.DataFrame,
    col_types_src: dict,
    col_types_tgt: dict,
) -> dict:
    """Monthly, quarterly, yearly aggregation comparison + rolling averages."""
    date_col = _find_date_col(src, col_types_src)
    if date_col is None or date_col not in tgt.columns:
        log.warning("No date column detected — skipping time analysis.")
        return {}

    num_cols = _find_numeric_cols(src, col_types_src)
    if not num_cols:
        return {}

    src2 = src[[date_col] + num_cols].copy()
    tgt2 = tgt[[date_col] + num_cols].copy()
    src2[date_col] = pd.to_datetime(src2[date_col], errors="coerce")
    tgt2[date_col] = pd.to_datetime(tgt2[date_col], errors="coerce")
    src2 = src2.dropna(subset=[date_col])
    tgt2 = tgt2.dropna(subset=[date_col])

    result = {}

    for freq, label in [("ME", "Monthly"), ("QE", "Quarterly"), ("YE", "Yearly")]:
        agg_src = src2.set_index(date_col).resample(freq)[num_cols].agg(["sum","count","nunique"])
        agg_tgt = tgt2.set_index(date_col).resample(freq)[num_cols].agg(["sum","count","nunique"])

        agg_src.columns = [f"SRC_{c}_{a}" for c, a in agg_src.columns]
        agg_tgt.columns = [f"TGT_{c}_{a}" for c, a in agg_tgt.columns]

        merged = agg_src.join(agg_tgt, how="outer").reset_index()
        merged.rename(columns={date_col: "Period"}, inplace=True)
        merged["Period"] = merged["Period"].dt.strftime(
            "%Y-%m" if freq == "ME" else
            "%Y-Q%q" if freq == "QE" else "%Y"
        )
        result[label] = merged

    # Rolling 3-month average (on first numeric column, sum)
    if "Monthly" in result and num_cols:
        col0 = num_cols[0]
        m = result["Monthly"].copy()
        src_sum_col = f"SRC_{col0}_sum"
        tgt_sum_col = f"TGT_{col0}_sum"
        if src_sum_col in m.columns:
            m["SRC_3M_Rolling_Avg"] = m[src_sum_col].rolling(3, min_periods=1).mean().round(2)
        if tgt_sum_col in m.columns:
            m["TGT_3M_Rolling_Avg"] = m[tgt_sum_col].rolling(3, min_periods=1).mean().round(2)
        result["Monthly"] = m

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 8 ─ ADVANCED DQ CHECKS
# ═══════════════════════════════════════════════════════════════════════════════

def advanced_dq_checks(
    src: pd.DataFrame,
    tgt: pd.DataFrame,
    col_types_src: dict,
) -> pd.DataFrame:
    """
    Per-column advanced checks:
    • Outlier detection (IQR + Z-score)
    • Distribution drift (KS test)
    • Consistency flags
    """
    common_num = [
        c for c, t in col_types_src.items()
        if t == "numeric" and c in src.columns and c in tgt.columns
    ]
    rows = []
    for col in common_num:
        sv = src[col].dropna()
        tv = tgt[col].dropna()
        if len(sv) < 10 or len(tv) < 10:
            continue

        # IQR outliers
        def iqr_outliers(s):
            q1, q3 = s.quantile(.25), s.quantile(.75)
            iqr = q3 - q1
            return int(((s < q1 - 1.5*iqr) | (s > q3 + 1.5*iqr)).sum())

        # Z-score outliers (|z| > 3)
        def zscore_outliers(s):
            return int((np.abs(scipy_stats.zscore(s)) > 3).sum())

        # KS test for distribution drift
        ks_stat, ks_p = scipy_stats.ks_2samp(sv.values, tv.values)

        src_iqr_out  = iqr_outliers(sv)
        tgt_iqr_out  = iqr_outliers(tv)
        src_z_out    = zscore_outliers(sv)
        tgt_z_out    = zscore_outliers(tv)

        rows.append({
            "Column"                  : col,
            "KS Statistic"            : round(ks_stat, 4),
            "KS p-value"              : round(ks_p, 4),
            "Distribution Drift"      : "⚠️ Drift Detected" if ks_p < 0.05 else "✅ Stable",
            "Source IQR Outliers"     : src_iqr_out,
            "Target IQR Outliers"     : tgt_iqr_out,
            "Source Z-Score Outliers" : src_z_out,
            "Target Z-Score Outliers" : tgt_z_out,
            "Source Min"              : round(float(sv.min()), 4),
            "Source Max"              : round(float(sv.max()), 4),
            "Target Min"              : round(float(tv.min()), 4),
            "Target Max"              : round(float(tv.max()), 4),
            "Range Δ"                 : round(
                (float(tv.max()) - float(tv.min())) -
                (float(sv.max()) - float(sv.min())), 4
            ),
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 9 ─ DQ SCORECARD
# ═══════════════════════════════════════════════════════════════════════════════

def compute_scorecard(
    profile_df: pd.DataFrame,
    mismatch_df: pd.DataFrame,
    row_stats: dict,
    schema_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Four dimensions: Completeness, Consistency, Accuracy, Validity.
    Score each 0–100; compute overall weighted DQ Score.
    """
    # Completeness: average fill-rate across columns
    if "Source Fill Rate %" in profile_df.columns:
        completeness = profile_df["Source Fill Rate %"].mean()
    else:
        completeness = 100.0

    # Consistency: schema match %
    schema_ok = (schema_df["Status"] == "✅ Both").sum()
    consistency = schema_ok / max(len(schema_df), 1) * 100

    # Accuracy: 1 - weighted mismatch %
    if "Mismatch %" in mismatch_df.columns and len(mismatch_df):
        accuracy = 100 - mismatch_df["Mismatch %"].mean()
    else:
        accuracy = 100.0

    # Validity: row duplicate ratio
    validity = 100 - row_stats.get("source_dup_pct", 0)

    overall = round(
        0.35 * completeness +
        0.25 * consistency   +
        0.30 * accuracy      +
        0.10 * validity, 2
    )

    def grade(score):
        if score >= 95: return "A — Excellent"
        if score >= 85: return "B — Good"
        if score >= 70: return "C — Acceptable"
        if score >= 50: return "D — Poor"
        return "F — Critical"

    rows = [
        {"Dimension": "Completeness", "Score": round(completeness, 2),
         "Weight": "35%", "Grade": grade(completeness),
         "Description": "Average fill-rate across all source columns"},
        {"Dimension": "Consistency",  "Score": round(consistency, 2),
         "Weight": "25%", "Grade": grade(consistency),
         "Description": "Schema / type alignment between source & target"},
        {"Dimension": "Accuracy",     "Score": round(accuracy, 2),
         "Weight": "30%", "Grade": grade(accuracy),
         "Description": "Cell-level value match between source & target"},
        {"Dimension": "Validity",     "Score": round(validity, 2),
         "Weight": "10%", "Grade": grade(validity),
         "Description": "Duplicate-free rows in source"},
        {"Dimension": "⭐ OVERALL DQ SCORE", "Score": overall,
         "Weight": "100%", "Grade": grade(overall),
         "Description": "Weighted composite score"},
    ]
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 10 ─ EXCEL REPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

# ── Colour palette ──────────────────────────────────────────────────────────
C = {
    "navy"     : "1F3864",
    "blue"     : "2E75B6",
    "lt_blue"  : "BDD7EE",
    "teal"     : "00B0A0",
    "green"    : "70AD47",
    "lt_green" : "E2EFDA",
    "amber"    : "FFC000",
    "lt_amber" : "FFF2CC",
    "red"      : "C00000",
    "lt_red"   : "FCE4D6",
    "white"    : "FFFFFF",
    "lt_grey"  : "F2F2F2",
    "mid_grey" : "D9D9D9",
    "dark_grey": "595959",
    "black"    : "000000",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Aptos Narrow" if bold else "Aptos Narrow")

def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _auto_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cv = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(cv))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def _write_header_row(ws, headers: list, row: int = 1,
                      bg=C["navy"], fg=C["white"], size=10):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font      = _font(bold=True, color=fg, size=size)
        cell.fill      = _fill(bg)
        cell.alignment = _align("center", wrap=True)
        cell.border    = _border()

def _write_data_rows(ws, df: pd.DataFrame, start_row: int = 2,
                     zebra=True, border=True):
    for ri, row_data in enumerate(df.itertuples(index=False), start_row):
        is_odd = (ri - start_row) % 2 == 0
        bg = C["lt_grey"] if (zebra and is_odd) else C["white"]
        for ci, val in enumerate(row_data, 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = _fill(bg)
            cell.alignment = _align("left" if isinstance(val, str) else "right",
                                    wrap=False)
            cell.font      = _font()
            if border:
                cell.border = _border()

def _sheet_title(ws, title: str, subtitle: str = ""):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font      = _font(bold=True, color=C["white"], size=14)
    c1.fill      = _fill(C["navy"])
    c1.alignment = _align("left", "center")
    if subtitle:
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font      = _font(color=C["dark_grey"], italic=True)
        c2.fill      = _fill(C["lt_blue"])
        c2.alignment = _align("left", "center")


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_summary_sheet(wb: Workbook, src: pd.DataFrame, tgt: pd.DataFrame,
                        row_stats: dict, mismatch_df: pd.DataFrame,
                        scorecard_df: pd.DataFrame, run_ts: str):
    ws = wb.create_sheet("📊 Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "  DATA QUALITY REPORT  —  Source vs Target"
    c.font      = Font(name="Aptos Narrow", bold=True, color=C["white"], size=16)
    c.fill      = _fill(C["navy"])
    c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value     = f"  Generated: {run_ts}   |   Source rows: {row_stats['source_rows']:,}   |   Target rows: {row_stats['target_rows']:,}"
    c2.font      = _font(italic=True, color=C["dark_grey"])
    c2.fill      = _fill(C["lt_blue"])
    c2.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 20

    # ── KPI tiles ─────────────────────────────────────────────────────────────
    kpis = [
        ("Source Rows",        f"{row_stats['source_rows']:,}",         C["blue"]),
        ("Target Rows",        f"{row_stats['target_rows']:,}",         C["blue"]),
        ("Row Delta",          f"{row_stats['row_delta']:+,}",          C["amber"]),
        ("Common Columns",     str(len([c for c in src.columns if c in tgt.columns])), C["teal"]),
        ("Total Mismatches",   f"{mismatch_df['Mismatch Count'].sum():,}" if len(mismatch_df) else "N/A", C["red"]),
        ("DQ Score",
         f"{scorecard_df[scorecard_df['Dimension'].str.contains('OVERALL')]['Score'].values[0]:.1f} / 100"
         if len(scorecard_df) else "N/A",
         C["green"]),
    ]

    for ti, (label, val, color) in enumerate(kpis):
        col_start = (ti % 3) * 2 + 1
        data_row  = 4 if ti < 3 else 8
        ws.merge_cells(
            start_row=data_row, start_column=col_start,
            end_row=data_row,   end_column=col_start + 1
        )
        lc = ws.cell(row=data_row - 1, column=col_start, value=label)
        lc.font      = _font(bold=True, color=C["white"], size=9)
        lc.fill      = _fill(color)
        lc.alignment = _align("center", "center")
        ws.merge_cells(
            start_row=data_row - 1, start_column=col_start,
            end_row=data_row - 1,   end_column=col_start + 1
        )
        vc = ws.cell(row=data_row, column=col_start, value=val)
        vc.font      = Font(name="Aptos Narrow", bold=True, size=20, color=color)
        vc.fill      = _fill(C["lt_grey"])
        vc.alignment = _align("center", "center")
        ws.row_dimensions[data_row].height = 35

    # ── Scorecard mini table ────────────────────────────────────────────────
    ws.cell(row=11, column=1, value="DQ SCORECARD SUMMARY").font = _font(bold=True, size=11, color=C["navy"])
    _write_header_row(ws, list(scorecard_df.columns), row=12, bg=C["blue"])
    _write_data_rows(ws, scorecard_df, start_row=13)

    # colour overall row
    for ci in range(1, len(scorecard_df.columns) + 1):
        ws.cell(row=13 + len(scorecard_df) - 1, column=ci).fill = _fill(C["lt_amber"])
        ws.cell(row=13 + len(scorecard_df) - 1, column=ci).font = _font(bold=True)

    # ── Mismatch top-10 table ───────────────────────────────────────────────
    if len(mismatch_df):
        top10 = mismatch_df.nlargest(10, "Mismatch %")[
            ["Column", "Mismatch Count", "Mismatch %", "DQ Status"]
        ].reset_index(drop=True)
        start = 13 + len(scorecard_df) + 3
        ws.cell(row=start - 1, column=1, value="TOP 10 COLUMNS BY MISMATCH %").font = _font(bold=True, size=11, color=C["navy"])
        _write_header_row(ws, list(top10.columns), row=start, bg=C["red"])
        _write_data_rows(ws, top10, start_row=start + 1)

    _auto_width(ws)
    ws.column_dimensions["A"].width = 28


def build_schema_sheet(wb: Workbook, schema_df: pd.DataFrame):
    ws = wb.create_sheet("🗂️ Schema Comparison")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Schema Comparison", "Column presence, position & data-type alignment")
    _write_header_row(ws, list(schema_df.columns), row=3, bg=C["blue"])
    _write_data_rows(ws, schema_df, start_row=4)

    # Conditional fill by status
    status_col = list(schema_df.columns).index("Status") + 1
    for ri in range(4, 4 + len(schema_df)):
        cell = ws.cell(row=ri, column=status_col)
        val  = str(cell.value)
        if "Missing" in val:
            ws.cell(row=ri, column=1).fill = _fill(C["lt_red"])
            cell.fill = _fill(C["lt_red"])
        elif "Type Mismatch" in val:
            ws.cell(row=ri, column=1).fill = _fill(C["lt_amber"])
            cell.fill = _fill(C["lt_amber"])
        else:
            cell.fill = _fill(C["lt_green"])

    _auto_width(ws)


def build_profile_sheet(wb: Workbook, profile_df: pd.DataFrame):
    ws = wb.create_sheet("📋 Column Profiling")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Column-Level Profiling", "Statistical profile for every common column (Source vs Target)")
    _write_header_row(ws, list(profile_df.columns), row=3, bg=C["teal"])
    _write_data_rows(ws, profile_df, start_row=4)

    # Highlight high null %
    null_col_idx = None
    for ci, h in enumerate(profile_df.columns, 1):
        if h == "Source Null %":
            null_col_idx = ci
            break
    if null_col_idx:
        for ri in range(4, 4 + len(profile_df)):
            cell = ws.cell(row=ri, column=null_col_idx)
            try:
                if float(cell.value or 0) > 20:
                    cell.fill = _fill(C["lt_red"])
                elif float(cell.value or 0) > 5:
                    cell.fill = _fill(C["lt_amber"])
            except (ValueError, TypeError):
                pass

    ws.freeze_panes = "B4"
    _auto_width(ws)


def build_mismatch_summary_sheet(wb: Workbook, mismatch_df: pd.DataFrame):
    ws = wb.create_sheet("❌ Mismatch Summary")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Cell-Level Mismatch Summary", "Column-wise match / mismatch analysis")
    _write_header_row(ws, list(mismatch_df.columns), row=3, bg=C["red"])
    _write_data_rows(ws, mismatch_df, start_row=4)

    # Red-scale on Mismatch %
    mm_col_idx = list(mismatch_df.columns).index("Mismatch %") + 1
    mm_col_letter = get_column_letter(mm_col_idx)
    last_row = 3 + len(mismatch_df)
    ws.conditional_formatting.add(
        f"{mm_col_letter}4:{mm_col_letter}{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color=C["lt_green"],
            mid_type="num",   mid_value=5,     mid_color=C["amber"],
            end_type="num",   end_value=100,   end_color=C["red"],
        )
    )
    ws.freeze_panes = "B4"
    _auto_width(ws)


def build_mismatch_samples_sheet(wb: Workbook, samples_df: pd.DataFrame):
    ws = wb.create_sheet("🔍 Mismatch Samples")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Detailed Mismatch Samples", "Row-level examples of source/target value differences")
    if len(samples_df) == 0:
        ws.cell(row=3, column=1, value="✅ No mismatches found — datasets are identical!").font = \
            _font(bold=True, color=C["green"], size=12)
        return
    _write_header_row(ws, list(samples_df.columns), row=3, bg=C["red"])
    _write_data_rows(ws, samples_df, start_row=4)

    mtype_col = list(samples_df.columns).index("Mismatch Type") + 1
    for ri in range(4, 4 + len(samples_df)):
        cell = ws.cell(row=ri, column=mtype_col)
        if "Value Mismatch" in str(cell.value):
            for ci in range(1, len(samples_df.columns) + 1):
                ws.cell(row=ri, column=ci).fill = _fill(C["lt_red"])
        elif "Null" in str(cell.value):
            for ci in range(1, len(samples_df.columns) + 1):
                ws.cell(row=ri, column=ci).fill = _fill(C["lt_amber"])

    _auto_width(ws)


def build_time_sheet(wb: Workbook, time_results: dict):
    if not time_results:
        ws = wb.create_sheet("📅 Time Analysis")
        ws.cell(row=1, column=1, value="No date column detected — time analysis not available.")
        return

    ws = wb.create_sheet("📅 Time Analysis")
    ws.sheet_view.showGridLines = False

    current_row = 1
    for period_label, df in time_results.items():
        if df is None or len(df) == 0:
            continue
        title_cell = ws.cell(row=current_row, column=1, value=f"▶ {period_label} Aggregation")
        title_cell.font = _font(bold=True, size=12, color=C["navy"])
        title_cell.fill = _fill(C["lt_blue"])
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=min(len(df.columns), 15)
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        cols = [c for c in df.columns if not df[c].isna().all()][:15]
        display_df = df[cols].copy()
        if "Period" in display_df.columns:
            display_df["Period"] = display_df["Period"].astype(str)

        _write_header_row(ws, list(display_df.columns), row=current_row, bg=C["blue"])
        current_row += 1

        _write_data_rows(ws, display_df, start_row=current_row)
        current_row += len(display_df) + 3

    _auto_width(ws)


def build_row_checks_sheet(wb: Workbook, row_stats: dict):
    ws = wb.create_sheet("📌 Row & Duplicate Checks")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Row-Level & Duplicate Analysis", "Count comparison, duplicates, key-based match summary")

    data = [
        ("Metric", "Source", "Target", "Delta / Note"),
        ("Total Rows",
         f"{row_stats['source_rows']:,}",
         f"{row_stats['target_rows']:,}",
         f"{row_stats['row_delta']:+,} ({row_stats['row_delta_pct']:+.2f}%)"),
        ("Duplicate Rows",
         f"{row_stats.get('source_duplicates', 0):,} ({row_stats.get('source_dup_pct',0):.2f}%)",
         f"{row_stats.get('target_duplicates', 0):,} ({row_stats.get('target_dup_pct',0):.2f}%)",
         ""),
    ]

    if "row_hash_match_pct" in row_stats:
        data.append((
            "Row Hash Match",
            f"{row_stats['row_hash_matches']:,}",
            f"{row_stats['row_hash_mismatches']:,} mismatches",
            f"{row_stats['row_hash_match_pct']:.2f}% match",
        ))
    if "keys_in_both" in row_stats:
        data += [
            ("Keys in Both",       str(row_stats.get("keys_in_both","")),   "", ""),
            ("Keys Missing in Tgt",str(row_stats.get("keys_in_src_not_tgt","")), "", ""),
            ("Keys Missing in Src",str(row_stats.get("keys_in_tgt_not_src","")), "", ""),
        ]

    for ri, row in enumerate(data, 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri == 3:
                cell.font   = _font(bold=True, color=C["white"])
                cell.fill   = _fill(C["navy"])
            else:
                cell.font   = _font()
                cell.fill   = _fill(C["lt_grey"] if ri % 2 == 0 else C["white"])
            cell.border    = _border()
            cell.alignment = _align("center")

    # Sample missing keys
    for key_type in ["sample_missing_in_tgt", "sample_missing_in_src"]:
        if key_type in row_stats and row_stats[key_type]:
            label = "Sample Keys Missing in Target:" if "tgt" in key_type else "Sample Keys Missing in Source:"
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value=label).font = _font(bold=True, color=C["red"])
            for i, k in enumerate(row_stats[key_type][:20], 1):
                ws.cell(row=r + i, column=1, value=str(k)).font = _font()

    _auto_width(ws)


def build_advanced_dq_sheet(wb: Workbook, adv_df: pd.DataFrame):
    ws = wb.create_sheet("🔬 Advanced DQ")
    ws.sheet_view.showGridLines = False
    _sheet_title(ws, "Advanced Data Quality Checks",
                 "KS distribution drift test, IQR / Z-score outlier detection per numeric column")
    if len(adv_df) == 0:
        ws.cell(row=3, column=1, value="No numeric columns available for advanced checks.").font = _font()
        return
    _write_header_row(ws, list(adv_df.columns), row=3, bg=C["teal"])
    _write_data_rows(ws, adv_df, start_row=4)

    drift_col = list(adv_df.columns).index("Distribution Drift") + 1
    for ri in range(4, 4 + len(adv_df)):
        cell = ws.cell(row=ri, column=drift_col)
        if "Drift" in str(cell.value):
            for ci in range(1, len(adv_df.columns) + 1):
                ws.cell(row=ri, column=ci).fill = _fill(C["lt_amber"])

    _auto_width(ws)


def build_scorecard_sheet(wb: Workbook, scorecard_df: pd.DataFrame):
    ws = wb.create_sheet("🏆 DQ Scorecard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "DATA QUALITY SCORECARD"
    c.font      = Font(name="Aptos Narrow", bold=True, color=C["white"], size=18)
    c.fill      = _fill(C["navy"])
    c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 50

    _write_header_row(ws, list(scorecard_df.columns), row=3, bg=C["blue"], size=11)

    grade_colors = {
        "A": C["lt_green"],
        "B": C["lt_green"],
        "C": C["lt_amber"],
        "D": C["lt_red"],
        "F": C["lt_red"],
    }

    for ri, row_data in enumerate(scorecard_df.itertuples(index=False), 4):
        is_overall = "OVERALL" in str(row_data[0])
        for ci, val in enumerate(row_data, 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = _align("center" if ci > 1 else "left", "center")
            if is_overall:
                cell.font = Font(name="Aptos Narrow", bold=True, size=12, color=C["navy"])
                cell.fill = _fill(C["lt_amber"])
            else:
                grade = str(row_data[3])[0] if len(str(row_data[3])) > 0 else ""
                bg    = grade_colors.get(grade, C["white"])
                cell.font = _font()
                cell.fill = _fill(bg)
        ws.row_dimensions[ri].height = 25

    # score bar chart
    try:
        chart = BarChart()
        chart.type    = "bar"
        chart.title   = "DQ Dimension Scores"
        chart.y_axis.title = "Score (0-100)"
        chart.style   = 10

        data_ref = Reference(ws, min_col=2, max_col=2,
                             min_row=3, max_row=3 + len(scorecard_df) - 2)
        cats_ref = Reference(ws, min_col=1, max_col=1,
                             min_row=4, max_row=3 + len(scorecard_df) - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, "A" + str(4 + len(scorecard_df) + 2))
    except Exception:
        pass

    _auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 11 ─ MAIN ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════════════

def run_dq_report(
    source_path: str,
    target_path: str,
    output_path: str = "DQ_Report.xlsx",
    key_col: Optional[str] = None,
    sample_mismatches: int = 50,
    chunksize: int = 500_000,
):
    """
    Full end-to-end DQ pipeline.

    Parameters
    ----------
    source_path        : path to source CSV
    target_path        : path to target CSV
    output_path        : output Excel file name
    key_col            : optional primary key column for key-based row matching
    sample_mismatches  : max mismatch samples per column in report
    chunksize          : rows per chunk for CSV reading
    """
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info("=" * 70)
    log.info("  ENTERPRISE DQ FRAMEWORK — Starting Run")
    log.info("=" * 70)

    # ── 1. Ingest ─────────────────────────────────────────────────────────────
    log.info("[1/9] Ingesting datasets …")
    src = load_csv(source_path, chunksize=chunksize, key_col=key_col)
    tgt = load_csv(target_path, chunksize=chunksize, key_col=key_col)

    # ── 2. Detect & coerce types ──────────────────────────────────────────────
    log.info("[2/9] Detecting column types …")
    ctype_src = detect_column_types(src)
    ctype_tgt = detect_column_types(tgt)
    src = coerce_types(src, ctype_src)
    tgt = coerce_types(tgt, ctype_tgt)

    # ── 3. Schema ─────────────────────────────────────────────────────────────
    log.info("[3/9] Comparing schemas …")
    schema_df = compare_schemas(src, tgt)

    # ── 4. Column profiling ───────────────────────────────────────────────────
    log.info("[4/9] Profiling columns …")
    profile_df = profile_all_columns(src, tgt, ctype_src, ctype_tgt)

    # ── 5. Cell-level comparison ──────────────────────────────────────────────
    log.info("[5/9] Cell-level comparison …")
    mismatch_df, samples_df = cell_level_comparison(
        src, tgt, sample_n=sample_mismatches, key_col=key_col
    )

    # ── 6. Row-level checks ───────────────────────────────────────────────────
    log.info("[6/9] Row-level checks …")
    row_stats = row_level_checks(src, tgt, key_col=key_col)

    # ── 7. Time-based analysis ────────────────────────────────────────────────
    log.info("[7/9] Time-based analysis …")
    time_results = time_based_analysis(src, tgt, ctype_src, ctype_tgt)

    # ── 8. Advanced DQ ───────────────────────────────────────────────────────
    log.info("[8/9] Advanced DQ checks (drift, outliers) …")
    adv_df = advanced_dq_checks(src, tgt, ctype_src)

    # ── 9. Scorecard ─────────────────────────────────────────────────────────
    log.info("[9/9] Computing DQ scorecard …")
    scorecard_df = compute_scorecard(profile_df, mismatch_df, row_stats, schema_df)

    # ── Build Excel ───────────────────────────────────────────────────────────
    log.info("Building Excel report …")
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    build_summary_sheet  (wb, src, tgt, row_stats, mismatch_df, scorecard_df, run_ts)
    build_schema_sheet   (wb, schema_df)
    build_profile_sheet  (wb, profile_df)
    build_mismatch_summary_sheet(wb, mismatch_df)
    build_mismatch_samples_sheet(wb, samples_df)
    build_time_sheet     (wb, time_results)
    build_row_checks_sheet(wb, row_stats)
    build_advanced_dq_sheet(wb, adv_df)
    build_scorecard_sheet(wb, scorecard_df)

    wb.save(output_path)
    log.info(f"✅  Report saved → {output_path}")
    log.info("=" * 70)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Enterprise Data Quality Framework — Source vs Target"
    )
    parser.add_argument("source",  help="Path to source CSV file")
    parser.add_argument("target",  help="Path to target CSV file")
    parser.add_argument("--output",  default="DQ_Report.xlsx",
                        help="Output Excel file path (default: DQ_Report.xlsx)")
    parser.add_argument("--key",     default=None,
                        help="Primary key column name (optional)")
    parser.add_argument("--samples", type=int, default=50,
                        help="Max mismatch sample rows per column (default: 50)")
    parser.add_argument("--chunksize", type=int, default=500_000,
                        help="CSV chunk size for large files (default: 500,000)")
    args = parser.parse_args()

    run_dq_report(
        source_path=args.source,
        target_path=args.target,
        output_path=args.output,
        key_col=args.key,
        sample_mismatches=args.samples,
        chunksize=args.chunksize,
    )
